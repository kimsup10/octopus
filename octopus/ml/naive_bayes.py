from octopus.sns import get_articles
from openpyxl import load_workbook


class naive:
    def __init__(self):
        self.articles = get_articles('yadoran_q')
        self.get_users_size = len(self.articles[0].liked_users)




    def set_naive_bayes(self):

        temp_text = str()
        first_col = 'A'
        second_col = 'B'
        temp_dir = str()
        temp_num = 1
        total_follower = 3 #임시

        wb = load_workbook("tests/data/T1.xlsx")
        ws = wb.active


        for i in range(0, len(self.articles)):
            temp_text = self.articles[i].text.split('#')

            for j in range(1, len(temp_text)):
                ws[first_col + str(temp_num)] = temp_text[j]
                ws[second_col + str(temp_num)] = len(self.articles[i].liked_users) / total_follower
                temp_num = temp_num + 1



        wb.save("tests/data/T1.xlsx")



    def cal_naive_bayes(self):

        total_follower = 3 # 임시

        words = []
        temp_words = str()
        words_num = str()

        fir_col = 'A'
        sec_col = 'B'
        numb = 1

        total_naive = 1.0
        words_naive = []

        count = 0


        wb = load_workbook("tests/data/T1.xlsx")
        ws = wb.active


        print("Insert the words what you write on the instagram\n")

        words_num = input('How many word, do you input? ')

        for i in range(0, int(words_num)):
            temp_words = input(str(i+1) + ': ')
            words.append(temp_words)
            temp_words = None
            words_naive.append(0)


        print(words)
        print(words_naive)



        for i in range(0, int(words_num)):
            while(True):
                if(ws[fir_col + str(numb)].value == None):
                    numb = 1
                    break
                elif((ws[fir_col + str(numb)].value).find(words[i]) != -1):
                    print(ws[fir_col + str(numb)].value)
                    if(words_naive[i] < float(ws[sec_col + str(numb)].value)):
                        words_naive[i] = float(ws[sec_col + str(numb)].value)

                numb = numb + 1


        print(words_naive)

        for i in range(0, int(words_num)):
            if(words_naive[i] != 0):
                total_naive = total_naive * words_naive[i]
                count = count + 1

        total_naive = total_naive ** (1.0/float(count))

        print("Total naive bayes is : " + str(total_naive))